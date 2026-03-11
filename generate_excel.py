import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import zipfile

# Define paths
base_dir = os.path.dirname(os.path.abspath(__file__))
input_excel_path = os.path.join(base_dir, "Mike-prepared.xlsx")
output_excel_path = os.path.join(base_dir, "telecom_sales_scoring_model.xlsx")
zip_path = os.path.join(base_dir, "telecom_sales_scoring_model.zip")

# Load input data
wb_in = openpyxl.load_workbook(input_excel_path, data_only=True)
ws_in = wb_in.active

# Find column indices
headers_in = [cell.value for cell in ws_in[1]]
col_idx = {name: idx for idx, name in enumerate(headers_in)}

# Create output workbook
wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "Sales_Data"

headers_out = [
    "区域",
    "客户经理数",
    "客户数",
    "规模以上企业数",
    "区域企业总数",
    "2024收入(万)",
    "2025收入(万)",
    "2025收入目标",
    "净增收入完成率",
    "收入增幅",
    "单位客户收入(万)",
    "人均产值(万)",
    "人均新增收入(万)",
    "客户渗透率",
    "人均服务客户数",
    "意向客户数",
    "潜在合作金额(万)",
    "实际签约率",
    "数字化转型签约(万)",
    "5G工厂申报",
    "智慧工厂申报",
    "标准ICT数",
    "规模得分(30)",
    "增长得分(30)",
    "效率得分(25)",
    "潜力得分(15)",
    "战略加分(10)",
    "综合得分(110)",
    "排名"
]

# Write headers
for col, h in enumerate(headers_out, start=1):
    cell = ws_out.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True)

# Write data and add formulas
data_rows = list(ws_in.iter_rows(min_row=2, values_only=True))
last_row = len(data_rows) + 1
if last_row < 2:
    last_row = 2

for r, row_data in enumerate(data_rows, start=2):
    # Retrieve Base Data
    region = row_data[col_idx.get('区县', 0)] if '区县' in col_idx else ""
    managers = row_data[col_idx.get('客户经理数', 0)] if '客户经理数' in col_idx else 0
    clients = row_data[col_idx.get('客户数', 0)] if '客户数' in col_idx else 0
    scale_ents = row_data[col_idx.get('区域内规模以上企业数（指2000万收入以上的企业)', 0)] if '区域内规模以上企业数（指2000万收入以上的企业)' in col_idx else 0
    non_clients = row_data[col_idx.get('区域内未成为客户的数量', 0)] if '区域内未成为客户的数量' in col_idx else 0
    
    rev_2024 = row_data[col_idx.get('区域内24年收入', 0)] if '区域内24年收入' in col_idx else 0
    rev_2025 = row_data[col_idx.get('2025年工业收入（万元）', 0)] if '2025年工业收入（万元）' in col_idx else 0
    tar_2025 = row_data[col_idx.get('2025年收入目标', 0)] if '2025年收入目标' in col_idx else 0
    
    net_growth_pct = row_data[col_idx.get('净增收入完成率', 0)] if '净增收入完成率' in col_idx else 0
    rev_growth_pct = row_data[col_idx.get('收入增幅', 0)] if '收入增幅' in col_idx else 0
    
    unit_client_rev = row_data[col_idx.get('单位客户收入（万元）', 0)] if '单位客户收入（万元）' in col_idx else 0
    per_capita_out = row_data[col_idx.get('人均产值（万元）', 0)] if '人均产值（万元）' in col_idx else 0
    
    intended_clients = row_data[col_idx.get('意向客户数量', 0)] if '意向客户数量' in col_idx else 0
    potential_amount = row_data[col_idx.get('潜在合作商机金额（万元）', 0)] if '潜在合作商机金额（万元）' in col_idx else 0
    actual_sign_rate = row_data[col_idx.get('按金额统计的实际签约率', 0)] if '按金额统计的实际签约率' in col_idx else 0
    
    digital_sign = row_data[col_idx.get('25年数字化转型签约金额（万元）', 0)] if '25年数字化转型签约金额（万元）' in col_idx else 0
    factory_5g = row_data[col_idx.get('5G工厂申报数', 0)] if '5G工厂申报数' in col_idx else 0
    smart_factory = row_data[col_idx.get('智慧工厂申报数', 0)] if '智慧工厂申报数' in col_idx else 0
    std_ict = row_data[col_idx.get('标准ICT数', 0)] if '标准ICT数' in col_idx else 0

    # Calculated metrics
    total_ents = (clients or 0) + (non_clients or 0)
    
    # Write values iteratively
    ws_out[f"A{r}"] = region              # 区域
    ws_out[f"B{r}"] = managers            # 客户经理数
    ws_out[f"C{r}"] = clients             # 客户数
    ws_out[f"D{r}"] = scale_ents          # 规模企业
    ws_out[f"E{r}"] = total_ents          # 区域企业总数
    ws_out[f"F{r}"] = rev_2024            # 2024收入
    ws_out[f"G{r}"] = rev_2025            # 2025收入
    ws_out[f"H{r}"] = tar_2025            # 2025目标
    ws_out[f"I{r}"] = net_growth_pct      # 净增完成率
    ws_out[f"J{r}"] = rev_growth_pct      # 收入增幅
    ws_out[f"K{r}"] = unit_client_rev     # 单位客户收入
    ws_out[f"L{r}"] = per_capita_out      # 人均产值
    
    # Formulas calculated in Excel directly
    ws_out[f"M{r}"] = f"=IF(B{r}=0,0,(G{r}-F{r})/B{r})"           # 人均新增收入 (G-F / B)
    ws_out[f"N{r}"] = f"=IF(E{r}=0,0,C{r}/E{r})"                  # 客户渗透率 (C/E)
    ws_out[f"O{r}"] = f"=IF(B{r}=0,0,C{r}/B{r})"                  # 人均服务客户数 (C/B)
    
    ws_out[f"P{r}"] = intended_clients    # 意向客户数
    ws_out[f"Q{r}"] = potential_amount    # 潜在合作金额
    ws_out[f"R{r}"] = actual_sign_rate    # 实际签约率
    
    ws_out[f"S{r}"] = digital_sign        # 数字化转型签约
    ws_out[f"T{r}"] = factory_5g          # 5G
    ws_out[f"U{r}"] = smart_factory       # 智慧
    ws_out[f"V{r}"] = std_ict             # 标ICT
    
    # Max Scoring calculations
    # Scale (30) = 2025 Rev (20) + Unit Client Rev (10)
    # Using simple proportions vs MAX logic as standard representation (Value / MAX * Weight)
    ws_out[f"W{r}"] = f"=(IF(MAX($G$2:$G${last_row})=0,0,G{r}/MAX($G$2:$G${last_row}))*20) + (IF(MAX($K$2:$K${last_row})=0,0,K{r}/MAX($K$2:$K${last_row}))*10)"
    
    # Growth (30) = Revenue Growth (15) + Net Growth Completion (15)
    ws_out[f"X{r}"] = f"=(IF(MAX($J$2:$J${last_row})=0,0,J{r}/MAX($J$2:$J${last_row}))*15) + (IF(MAX($I$2:$I${last_row})=0,0,I{r}/MAX($I$2:$I${last_row}))*15)"
    
    # Efficiency (25) = Per Capita Output (10) + Client Def (10) + Clients/Mgr (5)
    ws_out[f"Y{r}"] = f"=(IF(MAX($L$2:$L${last_row})=0,0,L{r}/MAX($L$2:$L${last_row}))*10) + (IF(MAX($N$2:$N${last_row})=0,0,N{r}/MAX($N$2:$N${last_row}))*10) + (IF(MAX($O$2:$O${last_row})=0,0,O{r}/MAX($O$2:$O${last_row}))*5)"
    
    # Potential (15) = Potential Amount (5) + Intended Clients (5) + Actual Sign Rate (5)
    ws_out[f"Z{r}"] = f"=(IF(MAX($Q$2:$Q${last_row})=0,0,Q{r}/MAX($Q$2:$Q${last_row}))*5) + (IF(MAX($P$2:$P${last_row})=0,0,P{r}/MAX($P$2:$P${last_row}))*5) + (IF(MAX($R$2:$R${last_row})=0,0,R{r}/MAX($R$2:$R${last_row}))*5)"
    
    # Strategic Bonus (10) = Digital Sign (+5) + 5G (+2) + Smart (+2) + ICT (+1)
    ws_out[f"AA{r}"] = f"=MIN(10, (IF(MAX($S$2:$S${last_row})=0,0,S{r}/MAX($S$2:$S${last_row}))*5) + (IF(MAX($T$2:$T${last_row})=0,0,T{r}/MAX($T$2:$T${last_row}))*2) + (IF(MAX($U$2:$U${last_row})=0,0,U{r}/MAX($U$2:$U${last_row}))*2) + (IF(MAX($V$2:$V${last_row})=0,0,V{r}/MAX($V$2:$V${last_row}))*1))"
    
    # Overall Score = Scale + Growth + Efficiency + Potential + Bonus
    ws_out[f"AB{r}"] = f"=SUM(W{r}:AA{r})"
    
    # Ranking
    ws_out[f"AC{r}"] = f"=RANK(AB{r}, $AB$2:$AB${last_row}, 0)"

# Adjust column widths
for i in range(1, len(headers_out)+1):
    ws_out.column_dimensions[get_column_letter(i)].width = 16

# Save Excel
wb_out.save(output_excel_path)

# Zip it
with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as z:
    z.write(output_excel_path, os.path.basename(output_excel_path))

print(f"Generated {output_excel_path} and {zip_path}")
